"""
main.py — Extractor de Facturas PDF (Fiscalidad Española)
Modelo de salida: base de datos relacional (una fila por línea de detalle)

Flujo principal:
  1. Recorre todos los PDFs en la carpeta `facturas_in/`.
  2. Extrae el texto de cada PDF con PyMuPDF.
  3. Aplica el filtro OCR: si hay menos de OCR_MIN_CHARS caracteres, emite Warning.
  4. Envía el texto a la API de Gemini solicitando output estructurado (Pydantic).
  5. Expande cada factura en N filas (una por concepto/servicio).
  6. Calcula CUOTA_IVA_LINEA, CUOTA_IRPF_LINEA y TOTAL_LINEA por fila.
  7. Exporta a `facturas_procesadas.xlsx`.

Uso:
  python main.py
"""

from __future__ import annotations
import os
os.environ.setdefault("PYTHONUTF8", "1")

import logging
import sys
import warnings
from pathlib import Path
from typing import List, Optional, Any

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Font, PatternFill
from google import genai
from google.genai import types
from pydantic import BaseModel, Field, ValidationError

import config

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
# Forzar UTF-8 en la consola de Windows para evitar UnicodeEncodeError
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# ESQUEMAS PYDANTIC
# ---------------------------------------------------------------------------

class LineaDetalle(BaseModel):
    """Representa un concepto o servicio individual dentro de la factura."""

    CONCEPTO_DETALLE: Optional[str] = Field(
        None,
        description="Descripción completa del concepto, servicio o producto facturado."
    )
    BASE_IMPONIBLE_LINEA: Optional[float] = Field(
        None,
        description="Importe neto (base imponible) de este concepto concreto, en euros."
    )
    TIPO_IVA_PORCENTAJE: Optional[float] = Field(
        None,
        description="El porcentaje de IVA aplicado a esta linea (ej. 21.0, 10.0, 4.0, 0.0)"
    )
    TIPO_IRPF_PORCENTAJE: Optional[float] = Field(
        None,
        description="El porcentaje de retencion de IRPF aplicado a esta linea, si procede (ej. 15.0, 7.0, 0.0)"
    )
    CATEGORIA_CONTABLE: Optional[str] = Field(
        None,
        description=(
            "Categoria contable/fiscal asignada automaticamente segun el tipo de concepto. "
            "Ejemplos: 'Servicios Profesionales', 'Suministros', 'Licencias Informaticas', "
            "'Alquiler', 'Publicidad y Marketing', 'Transporte y Mensajeria', "
            "'Material de Oficina', 'Formacion', 'Mantenimiento y Reparaciones', "
            "'Honorarios', 'Otros Gastos'. Elige la mas adecuada segun el texto del concepto."
        )
    )


class Factura(BaseModel):
    """Cabecera de la factura + lista de lineas de detalle."""

    # --- Datos de cabecera ---
    NUM_FACT: Optional[str] = Field(None, description="Numero de factura")
    FECHA_EMISION: Optional[str] = Field(None, description="Fecha de emision (DD/MM/AAAA)")
    IVA: Optional[float] = Field(None, description="Porcentaje de IVA aplicado (ej. 21.0)")
    IRPF: Optional[float] = Field(None, description="Porcentaje de IRPF retenido (ej. 15.0)")
    RECEPTOR: Optional[str] = Field(None, description="Nombre o razon social del receptor")
    NIF_RECEPTOR: Optional[str] = Field(None, description="NIF/CIF del receptor (string)")
    DIRECCION_RECEPTOR: Optional[str] = Field(None, description="Direccion fiscal del receptor")
    CP_RECEPTOR: Optional[str] = Field(None, description="Codigo postal del receptor")
    POBLACION_RECEPTOR: Optional[str] = Field(None, description="Ciudad o poblacion del receptor")
    EMAIL_RECEPTOR: Optional[str] = Field(None, description="Email del receptor si aparece")
    EMITIDA: Optional[str] = Field(None, description="Nombre/razon social del emisor de la factura")
    PAGADA: Optional[str] = Field(None, description="Estado de pago: 'Si', 'No' o 'Desconocido'")

    # --- Lineas de detalle ---
    SERVICIOS: List[LineaDetalle] = Field(
        default_factory=list,
        description=(
            "Lista de todos los conceptos o servicios facturados. "
            "Cada elemento es una linea independiente con su descripcion, "
            "base imponible y categoria contable."
        )
    )


# ---------------------------------------------------------------------------
# COLUMNAS FINALES DEL EXCEL (orden exacto requerido)
# ---------------------------------------------------------------------------
COLUMNAS_EXCEL: list[str] = [
    "NUM_FACT",
    "FECHA_EMISION",
    "RECEPTOR",
    "NIF_RECEPTOR",
    "DIRECCION_RECEPTOR",
    "CP_RECEPTOR",
    "POBLACION_RECEPTOR",
    "EMAIL_RECEPTOR",
    "CATEGORIA_CONTABLE",
    "CONCEPTO_DETALLE",
    "BASE_IMPONIBLE_LINEA",
    "TIPO_IVA_PORCENTAJE",
    "CUOTA_IVA_LINEA",
    "TIPO_IRPF_PORCENTAJE",
    "CUOTA_IRPF_LINEA",
    "TOTAL_LINEA",
    "EMITIDA",
    "PAGADA",
]


# ---------------------------------------------------------------------------
# PROMPT MAESTRO
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """
Eres un asistente experto en contabilidad y fiscalidad espanola.
Tu unica tarea es extraer datos de facturas y devolverlos en formato JSON estructurado.

Reglas absolutas:
- Solo devuelves el JSON. Sin explicaciones, sin markdown, sin texto extra.
- Respeta exactamente los nombres de campo del esquema proporcionado.
- Los campos de precio y porcentaje deben ser numeros flotantes (ej. 21.0, 150.50).
- NIF, IBAN y codigos postales deben ser strings aunque sean numericos.
- Si un campo no aparece en la factura, devuelve null.
- Para PAGADA usa: "Si", "No" o "Desconocido".
- Las fechas deben tener formato DD/MM/AAAA siempre que sea posible.
- EMITIDA es el nombre del emisor (quien envia la factura), RECEPTOR es quien la recibe.

Para el campo SERVICIOS:
- Crea una entrada por cada linea de concepto o servicio distinto que aparezca en la factura.
- Extrae el TIPO_IVA_PORCENTAJE (ej. 21.0) y TIPO_IRPF_PORCENTAJE (ej. 15.0) para la linea de detalle (escribe solo el numero, no el signo %).
- Si solo hay un concepto global, crea una unica entrada.
- En BASE_IMPONIBLE_LINEA pon el importe neto de ese concepto concreto (sin IVA ni IRPF).
- En CATEGORIA_CONTABLE asigna la categoria contable mas adecuada segun el texto del concepto.
  Usa categorias coherentes como: 'Servicios Profesionales', 'Honorarios', 'Licencias Informaticas',
  'Suministros', 'Alquiler', 'Publicidad y Marketing', 'Transporte y Mensajeria',
  'Material de Oficina', 'Formacion', 'Mantenimiento y Reparaciones', 'Otros Gastos'.
- NUNCA dejes la lista SERVICIOS vacia. Si no hay desglose, pon un unico elemento con el total neto.
"""


# ---------------------------------------------------------------------------
# EXTRACCION DE TEXTO PDF
# ---------------------------------------------------------------------------
def extraer_texto_pdf(ruta_pdf: Path) -> str:
    """
    Extrae el texto completo de un PDF usando PyMuPDF.
    Devuelve una cadena vacia si ocurre cualquier error.
    """
    texto: str = ""
    try:
        with fitz.open(str(ruta_pdf)) as doc:
            for pagina in doc:
                texto += pagina.get_text()
    except Exception as exc:
        log.error("Error leyendo '%s': %s", ruta_pdf.name, exc)
    return texto.strip()


# ---------------------------------------------------------------------------
# LLAMADA A GEMINI
# ---------------------------------------------------------------------------
def extraer_datos_con_gemini(
    client: genai.Client,
    texto_factura: str,
    nombre_archivo: str,
) -> Optional[Factura]:
    """
    Envia el texto de la factura a Gemini y devuelve un objeto Factura validado.
    Retorna None si la API falla o la respuesta no pasa la validacion Pydantic.
    """
    prompt_usuario = (
        f"Extrae los datos de la siguiente factura espanola:\n\n"
        f"=== INICIO FACTURA: {nombre_archivo} ===\n"
        f"{texto_factura}\n"
        f"=== FIN FACTURA ==="
    )

    try:
        response = client.models.generate_content(
            model=config.GEMINI_MODEL,
            contents=prompt_usuario,
            config=types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
                response_mime_type="application/json",
                response_schema=Factura,
                temperature=0.0,  # Maxima determinismo
            ),
        )

        resultado = response.parsed
        if resultado is None:
            raise ValueError("La API devolvio una respuesta vacia o sin parsear.")

        return resultado  # type: ignore[return-value]

    except ValidationError as ve:
        log.error("Validacion Pydantic fallida para '%s': %s", nombre_archivo, ve)
        return None
    except Exception as exc:
        log.error("Error en la llamada a Gemini para '%s': %s", nombre_archivo, exc)
        return None


# ---------------------------------------------------------------------------
# EXPANSION DE FACTURA A FILAS RELACIONALES
# ---------------------------------------------------------------------------
def expandir_factura_a_filas(factura: Factura) -> list[dict]:
    """
    Transforma un objeto Factura en una lista de dicts, uno por cada
    linea de detalle (SERVICIO). Los datos de cabecera se repiten en cada fila.
    Calcula CUOTA_IVA_LINEA, CUOTA_IRPF_LINEA y TOTAL_LINEA por fila.
    """
    iva_global: float = factura.IVA or 0.0
    irpf_global: float = factura.IRPF or 0.0

    # Datos de cabecera comunes a todas las filas
    cabecera: dict[str, Any] = {
        "NUM_FACT":           factura.NUM_FACT,
        "FECHA_EMISION":      factura.FECHA_EMISION,
        "RECEPTOR":           factura.RECEPTOR,
        "NIF_RECEPTOR":       factura.NIF_RECEPTOR,
        "DIRECCION_RECEPTOR": factura.DIRECCION_RECEPTOR,
        "CP_RECEPTOR":        factura.CP_RECEPTOR,
        "POBLACION_RECEPTOR": factura.POBLACION_RECEPTOR,
        "EMAIL_RECEPTOR":     factura.EMAIL_RECEPTOR,
        "EMITIDA":            factura.EMITIDA,
        "PAGADA":             factura.PAGADA,
    }

    filas: list[dict] = []
    servicios = factura.SERVICIOS or []

    # Si Gemini no devolvio ninguna linea, generamos una fila vacia de placeholder
    if not servicios:
        log.warning(
            "Factura '%s' no tiene lineas de detalle. Se genera fila vacia.",
            factura.NUM_FACT or "SIN_NUM"
        )
        fila = {**cabecera}
        fila["CATEGORIA_CONTABLE"]   = None
        fila["CONCEPTO_DETALLE"]     = None
        fila["BASE_IMPONIBLE_LINEA"] = None
        fila["TIPO_IVA_PORCENTAJE"]  = None
        fila["CUOTA_IVA_LINEA"]      = None
        fila["TIPO_IRPF_PORCENTAJE"] = None
        fila["CUOTA_IRPF_LINEA"]     = None
        fila["TOTAL_LINEA"]          = None
        filas.append(fila)
        return filas

    for linea in servicios:
        base: Optional[float] = linea.BASE_IMPONIBLE_LINEA

        # Calculos fiscales por linea
        if base is not None:
            obj_iva_pct = linea.TIPO_IVA_PORCENTAJE if linea.TIPO_IVA_PORCENTAJE is not None else iva_global
            obj_irpf_pct = linea.TIPO_IRPF_PORCENTAJE if linea.TIPO_IRPF_PORCENTAJE is not None else irpf_global
            
            cuota_iva  = round(base * obj_iva_pct  / 100, 2)
            cuota_irpf = round(base * obj_irpf_pct / 100, 2)
            total      = round(base + cuota_iva - cuota_irpf, 2)
        else:
            obj_iva_pct = None
            obj_irpf_pct = None
            cuota_iva  = None
            cuota_irpf = None
            total      = None

        fila = {**cabecera}
        fila["CATEGORIA_CONTABLE"]   = linea.CATEGORIA_CONTABLE
        fila["CONCEPTO_DETALLE"]     = linea.CONCEPTO_DETALLE
        fila["BASE_IMPONIBLE_LINEA"] = base
        fila["TIPO_IVA_PORCENTAJE"]  = obj_iva_pct
        fila["CUOTA_IVA_LINEA"]      = cuota_iva
        fila["TIPO_IRPF_PORCENTAJE"] = obj_irpf_pct
        fila["CUOTA_IRPF_LINEA"]     = cuota_irpf
        fila["TOTAL_LINEA"]          = total
        filas.append(fila)

    return filas


# ---------------------------------------------------------------------------
# EXPORTACION A EXCEL
# ---------------------------------------------------------------------------
def exportar_a_excel(filas: list[dict], ruta_salida: Path) -> None:
    """
    Serializa la lista de filas relacionales a un archivo .xlsx con formato.
    """
    if not filas:
        log.warning("No hay datos nuevos para exportar.")
        return

    df_nuevo = pd.DataFrame(filas, columns=COLUMNAS_EXCEL)

    if ruta_salida.exists():
        try:
            df_existente = pd.read_excel(str(ruta_salida), engine="openpyxl")
            df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)
            log.info("Archivo Excel existente encontrado. Se anaden %d filas nuevas.", len(df_nuevo))
        except Exception as exc:
            log.warning("Error leyendo el Excel existente '%s'. Se creara uno nuevo. Error: %s", ruta_salida, exc)
            df_final = df_nuevo
    else:
        df_final = df_nuevo

    try:
        with pd.ExcelWriter(str(ruta_salida), engine="openpyxl") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Facturas")

            ws = writer.sheets["Facturas"]

            # --- Estilos de Cabecera Visual ---
            # 1. Fijar la primera fila
            ws.freeze_panes = "A2"

            # 2. Estilos para los titulos (Negrita y Fondo Azul)
            fuente_cabecera = Font(bold=True, color="FFFFFF")
            relleno_cabecera = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")

            # La primera fila contiene los encabezados (usar max_column para cubrir todas las columnas reales)
            for col_idx in range(1, ws.max_column + 1):
                celda = ws.cell(row=1, column=col_idx)
                celda.font = fuente_cabecera
                celda.fill = relleno_cabecera

            # 3. Aplicar Filtro Automatico a toda la tabla
            rango_tabla = f"A1:{ws.cell(row=len(df_final)+1, column=len(COLUMNAS_EXCEL)).coordinate}"
            ws.auto_filter.ref = rango_tabla

            # 4. Ajuste automatico de ancho de columnas
            for col_cells in ws.columns:
                max_len = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in col_cells
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        log.info(
            "[OK] Excel guardado en: %s  (%d fila/s de detalle)", ruta_salida, len(df_final)
        )
    except PermissionError:
        log.error(
            "\n[ERROR FATAL] No se puede guardar el Excel porque esta ABERTO.\n"
            "-----------------------------------------------------------\n"
            "👉 Por favor, CIERRA el archivo Excel (%s) y VUELVE A EJECUTAR el programa.\n"
            "-----------------------------------------------------------",
            ruta_salida.name
        )
        sys.exit(1)
    except Exception as exc:
        log.error("Error al escribir el Excel '%s': %s", ruta_salida, exc)
        raise


# ---------------------------------------------------------------------------
# MARCAR COMO PROCESADAS
# ---------------------------------------------------------------------------
def marcar_como_procesadas(pdfs: list[Path]) -> None:
    for p in pdfs:
        nuevo_nombre = f"[PROCESADA]_{p.name}"
        nueva_ruta = p.with_name(nuevo_nombre)
        try:
            p.rename(nueva_ruta)
            log.info("  [OK] PDF marcado como procesado: %s", nuevo_nombre)
        except Exception as exc:
            log.error("  [ERROR] No se pudo renombrar el archivo '%s': %s", p.name, exc)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main() -> None:
    log.info("=" * 60)
    log.info("  EXTRACTOR DE FACTURAS PDF - Fiscalidad Espanola")
    log.info("  Modelo: Relacional por Linea de Detalle")
    log.info("=" * 60)

    # --- Validacion de la API Key ---
    if config.GEMINI_API_KEY == "TU_API_KEY_AQUI" or not config.GEMINI_API_KEY:
        log.error(
            "[ERROR] No has configurado tu GEMINI_API_KEY.\n"
            "   Edita config.py o define la variable de entorno GEMINI_API_KEY."
        )
        sys.exit(1)

    # --- Inicializar cliente Gemini ---
    try:
        client = genai.Client(api_key=config.GEMINI_API_KEY)
        log.info("Cliente Gemini inicializado. Modelo: %s", config.GEMINI_MODEL)
    except Exception as exc:
        log.error("No se pudo inicializar el cliente de Gemini: %s", exc)
        sys.exit(1)

    # --- Verificar carpeta de entrada ---
    carpeta_entrada = Path(config.FACTURAS_DIR)
    if not carpeta_entrada.exists():
        log.info("La carpeta '%s' no existe. Creandola...", config.FACTURAS_DIR)
        carpeta_entrada.mkdir(parents=True, exist_ok=True)
        log.warning(
            "[AVISO] Carpeta '%s' creada. Coloca tus PDFs y vuelve a ejecutar.",
            config.FACTURAS_DIR,
        )
        sys.exit(0)

    pdfs_all = sorted(carpeta_entrada.glob("*.pdf"))
    pdfs = [p for p in pdfs_all if not p.name.startswith("[PROCESADA]")]
    omitidas = len(pdfs_all) - len(pdfs)
    
    if omitidas > 0:
        log.info("Ignorando %d factura(s) que ya estaban procesadas ('[PROCESADA]_...').", omitidas)

    if not pdfs:
        log.warning("No se encontraron archivos PDF en '%s'.", config.FACTURAS_DIR)
        sys.exit(0)

    log.info("PDFs encontrados: %d", len(pdfs))

    # --- Procesamiento ---
    todas_las_filas: list[dict] = []
    skipped_ocr: list[str] = []
    failed_api: list[str] = []
    pdfs_ok: list[Path] = []
    facturas_ok: int = 0

    for idx, ruta_pdf in enumerate(pdfs, start=1):
        log.info("-" * 50)
        log.info("[%d/%d] Procesando: %s", idx, len(pdfs), ruta_pdf.name)

        # 1. Extraer texto
        texto = extraer_texto_pdf(ruta_pdf)

        # 2. Filtro OCR
        if len(texto) < config.OCR_MIN_CHARS:
            warnings.warn(
                f"\n[WARNING OCR] El archivo '{ruta_pdf.name}' no tiene formato OCR "
                f"(texto extraido: {len(texto)} caracter/es). "
                f"Parece una imagen plana. Subelo en formato OCR.",
                UserWarning,
                stacklevel=1,
            )
            log.warning(
                "SKIPPED '%s' - texto insuficiente (%d chars < %d minimo).",
                ruta_pdf.name,
                len(texto),
                config.OCR_MIN_CHARS,
            )
            skipped_ocr.append(ruta_pdf.name)
            continue

        log.info("  Texto extraido: %d caracteres. Enviando a Gemini...", len(texto))

        # 3. Llamar a Gemini
        factura_obj = extraer_datos_con_gemini(client, texto, ruta_pdf.name)

        if factura_obj is None:
            log.error("  [ERROR] Fallo la extraccion para '%s'. Se omite.", ruta_pdf.name)
            failed_api.append(ruta_pdf.name)
            continue

        # 4. Expandir a filas relacionales
        filas = expandir_factura_a_filas(factura_obj)
        todas_las_filas.extend(filas)
        pdfs_ok.append(ruta_pdf)
        facturas_ok += 1

        log.info(
            "  [OK] Factura extraida: NUM_FACT=%s | %d linea/s de detalle",
            factura_obj.NUM_FACT or "-",
            len(filas),
        )

    # --- Exportar a Excel ---
    log.info("=" * 60)
    ruta_excel = Path(config.OUTPUT_FILE)
    exportar_a_excel(todas_las_filas, ruta_excel)

    # --- Renombrar PDFs procesados exitosamente ---
    if pdfs_ok:
        log.info("=" * 60)
        log.info("  Renombrando archivos procesados...")
        marcar_como_procesadas(pdfs_ok)

    # --- Resumen final ---
    log.info("")
    log.info("--- RESUMEN ---")
    log.info("  Facturas procesadas OK   : %d", facturas_ok)
    log.info("  Filas de detalle totales : %d", len(todas_las_filas))
    log.info("  Omitidas por falta OCR   : %d", len(skipped_ocr))
    log.info("  Fallidas por error API   : %d", len(failed_api))

    if skipped_ocr:
        log.warning("  Archivos sin OCR: %s", ", ".join(skipped_ocr))
    if failed_api:
        log.error("  Archivos con error API: %s", ", ".join(failed_api))

    log.info("=" * 60)


if __name__ == "__main__":
    main()
